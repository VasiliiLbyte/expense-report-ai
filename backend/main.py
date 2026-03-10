import os
import io
import json
import base64
from typing import List

from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv("backend/.env")

app = FastAPI(title="Expense Report AI")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

SYSTEM_PROMPT = """Ты — эксперт по распознаванию российских кассовых чеков.
Извлеки из изображения чека следующие поля и верни строго в JSON формате:
{
  "date": "дата в формате ДД.ММ.ГГГГ",
  "organization": "наименование организации/магазина",
  "inn": "ИНН организации (только цифры, или null если не найден)",
  "items": [
    {
      "name": "наименование товара/услуги",
      "amount_without_vat": "сумма без НДС (число)",
      "vat": "сумма НДС (число, 0 если не облагается)",
      "total": "итоговая сумма (число)"
    }
  ],
  "total_without_vat": "итого без НДС (число)",
  "total_vat": "итого НДС (число)",
  "total": "итого к оплате (число)"
}
Если поле не найдено — ставь null. Отвечай только JSON, без пояснений.
"""


def encode_image(file_bytes: bytes) -> str:
    return base64.b64encode(file_bytes).decode("utf-8")


def extract_receipt_data(image_bytes: bytes, filename: str) -> dict:
    base64_image = encode_image(image_bytes)
    ext = filename.lower().split(".")[-1]

    if ext in ["jpg", "jpeg"]:
        media_type = "image/jpeg"
    elif ext == "png":
        media_type = "image/png"
    elif ext == "webp":
        media_type = "image/webp"
    else:
        media_type = "image/jpeg"

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{media_type};base64,{base64_image}",
                            "detail": "high"
                        }
                    },
                    {
                        "type": "text",
                        "text": "Распознай чек и верни JSON"
                    }
                ]
            }
        ],
        max_tokens=2000,
        temperature=0
    )

    raw = response.choices[0].message.content.strip()

    if raw.startswith("```"):
        raw = raw.replace("```json", "").replace("```", "").strip()

    return json.loads(raw)


def build_preview_rows(receipts_data: List[dict]) -> List[dict]:
    rows = []

    for receipt_index, receipt in enumerate(receipts_data, start=1):
        items = receipt.get("items", [])

        if not items:
            items = [{
                "name": "—",
                "amount_without_vat": receipt.get("total_without_vat", 0),
                "vat": receipt.get("total_vat", 0),
                "total": receipt.get("total", 0),
            }]

        for item in items:
            rows.append({
                "receipt_no": receipt_index,
                "date": receipt.get("date", ""),
                "organization": receipt.get("organization", ""),
                "inn": receipt.get("inn", ""),
                "item_name": item.get("name", ""),
                "amount_without_vat": float(item.get("amount_without_vat") or 0),
                "vat": float(item.get("vat") or 0),
                "total": float(item.get("total") or 0),
                "filename": receipt.get("filename", ""),
            })

    return rows


def create_excel(receipts_data: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Авансовый отчёт"

    headers = [
        "№",
        "Дата",
        "Организация",
        "ИНН",
        "Наименование товара/услуги",
        "Сумма без НДС",
        "НДС",
        "Итого",
        "Файл",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "АВАНСОВЫЙ ОТЧЁТ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths =[2][3][4][5][6][7][8]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    row_num = 4

    for receipt_index, receipt in enumerate(receipts_data, start=1):
        items = receipt.get("items", [])

        if not items:
            items = [{
                "name": "—",
                "amount_without_vat": receipt.get("total_without_vat", 0),
                "vat": receipt.get("total_vat", 0),
                "total": receipt.get("total", 0),
            }]

        for item in items:
            values = [
                receipt_index,
                receipt.get("date", ""),
                receipt.get("organization", ""),
                receipt.get("inn", ""),
                item.get("name", ""),
                float(item.get("amount_without_vat") or 0),
                float(item.get("vat") or 0),
                float(item.get("total") or 0),
                receipt.get("filename", ""),
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if col in [6, 7, 8]:
                        cell.number_format = "#,##0.00"


            row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.post("/api/preview-receipts")
async def preview_receipts(files: List[UploadFile] = File(...)):
    receipts_data = []

    for file in files:
        try:
            content = await file.read()
            data = extract_receipt_data(content, file.filename)
            data["filename"] = file.filename
            receipts_data.append(data)
        except Exception as e:
            receipts_data.append({
                "date": "",
                "organization": "Ошибка распознавания",
                "inn": "",
                "items": [{
                    "name": str(e),
                    "amount_without_vat": 0,
                    "vat": 0,
                    "total": 0,
                }],
                "filename": file.filename,
            })

    rows = build_preview_rows(receipts_data)
    return JSONResponse({"rows": rows})


@app.post("/api/process-receipts")
async def process_receipts(files: List[UploadFile] = File(...)):
    receipts_data = []

    for file in files:
        try:
            content = await file.read()
            data = extract_receipt_data(content, file.filename)
            data["filename"] = file.filename
            receipts_data.append(data)
        except Exception as e:
            receipts_data.append({
                "date": "",
                "organization": "Ошибка распознавания",
                "inn": "",
                "items": [{
                    "name": str(e),
                    "amount_without_vat": 0,
                    "vat": 0,
                    "total": 0,
                }],
                "filename": file.filename,
            })

    excel_bytes = create_excel(receipts_data)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="avansoviy_otchet.xlsx"'
        },
    )
